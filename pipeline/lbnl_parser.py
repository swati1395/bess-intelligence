"""
LBNL Queued Up Parser — Module 4

Reads the LBNL "Queued Up" interconnection-queue Excel workbook, filters for
storage projects (any of type_1/type_2/type_3 in {Battery, Other Storage}) in
ERCOT and CAISO, and loads the structured rows into pipeline_projects in
data/intel.db.

Source dataset (annual release): https://emp.lbl.gov/queues
Drop the workbook into data/ (filename like lbnl_queued_up_YYYY.xlsx) before
running this parser.

NOTE ON UNITS: The LBNL dataset reports nameplate **power capacity in MW** —
there is no duration column, so true energy capacity (MWh / GWh) cannot be
computed from this file alone. This parser stores and reports MW (and GW for
aggregates). To estimate GWh, multiply by an assumed duration (industry
default for utility-scale BESS in ERCOT/CAISO is 2-4 hours).

Run: python3 pipeline/lbnl_parser.py [path/to/lbnl_queued_up_YYYY.xlsx]
"""

import argparse
import datetime as dt
import glob
import os
import sqlite3
import sys
from collections import defaultdict

import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from database.setup_db import DB_PATH, init_db  # noqa: E402

DATA_DIR = os.path.join(PROJECT_ROOT, "data")
SHEET_NAME = "03. Complete Queue Data"
HEADER_ROW = 2

TARGET_ISOS = {"ERCOT", "CAISO"}
STORAGE_TYPES = {"battery", "other storage"}

COLUMNS = [
    "q_id", "q_status", "q_date", "prop_date", "on_date", "wd_date", "ia_date",
    "IA_phase_raw", "IA_phase_clean", "county", "state", "fips_code",
    "poi_name", "region", "project_name", "utility", "entity", "developer",
    "cluster", "service", "project_type", "type_1", "type_2", "type_3",
    "type_clean", "mw_1", "mw_2", "mw_3", "q_year", "prop_year",
]


def auto_locate_workbook() -> str:
    candidates = sorted(glob.glob(os.path.join(DATA_DIR, "lbnl_queued_up_*.xlsx")), reverse=True)
    if not candidates:
        sys.exit(f"No LBNL workbook found in {DATA_DIR}/ — expected lbnl_queued_up_*.xlsx. "
                 f"Download from https://emp.lbl.gov/queues and place in data/.")
    return candidates[0]


def to_iso_date(value) -> str:
    """Convert openpyxl datetime/date cell to ISO 8601 date string. Empty/None → ''."""
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.date().isoformat() if isinstance(value, dt.datetime) else value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s  # leave as-is if unparseable


def to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def is_storage_type(t) -> bool:
    return t is not None and str(t).strip().lower() in STORAGE_TYPES


def parse_row(row: tuple, hi: dict) -> dict:
    """Pull fields from the row tuple via the header index. Returns dict suitable for insert."""
    g = lambda key: row[hi[key]] if key in hi else None
    types = [to_str(g("type_1")), to_str(g("type_2")), to_str(g("type_3"))]
    mws = [to_float(g("mw_1")), to_float(g("mw_2")), to_float(g("mw_3"))]

    storage_mw = sum(mw for t, mw in zip(types, mws) if mw is not None and is_storage_type(t))
    total_mw = sum(mw for mw in mws if mw is not None)
    has_storage = any(is_storage_type(t) for t in types)
    has_non_storage = any(t is not None and not is_storage_type(t) for t in types)

    return {
        "q_id": to_str(g("q_id")) if not isinstance(g("q_id"), (dt.datetime, dt.date)) else None,
        "iso_market": to_str(g("entity")),
        "q_status": to_str(g("q_status")),
        "q_date": to_iso_date(g("q_date")),
        "prop_date": to_iso_date(g("prop_date")),
        "on_date": to_iso_date(g("on_date")),
        "wd_date": to_iso_date(g("wd_date")),
        "ia_date": to_iso_date(g("ia_date")),
        "ia_phase": to_str(g("IA_phase_clean")) or to_str(g("IA_phase_raw")),
        "state": to_str(g("state")),
        "county": to_str(g("county")),
        "project_name": to_str(g("project_name")),
        "utility": to_str(g("utility")),
        "developer": to_str(g("developer")),
        "cluster": to_str(g("cluster")),
        "service": to_str(g("service")),
        "project_type": to_str(g("project_type")),
        "type_clean": to_str(g("type_clean")),
        "type_1": types[0],
        "type_2": types[1],
        "type_3": types[2],
        "mw_1": mws[0],
        "mw_2": mws[1],
        "mw_3": mws[2],
        "storage_mw": storage_mw if storage_mw else None,
        "total_mw": total_mw if total_mw else None,
        "is_hybrid": 1 if (has_storage and has_non_storage) else 0,
        "_has_storage": has_storage,
    }


def load_workbook_rows(path: str):
    print(f"Opening workbook: {path} ({os.path.getsize(path)/1024/1024:.1f} MB)")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames[:5]}…")
    ws = wb[SHEET_NAME]
    row_iter = ws.iter_rows(min_row=HEADER_ROW, values_only=True)
    headers = next(row_iter)
    hi = {h: i for i, h in enumerate(headers) if h}
    missing = [c for c in COLUMNS if c not in hi]
    if missing:
        print(f"  Warning: missing expected columns: {missing}", file=sys.stderr)
    return wb, row_iter, hi


def import_workbook(conn: sqlite3.Connection, path: str) -> dict:
    source_file = os.path.basename(path)
    wb, row_iter, hi = load_workbook_rows(path)

    cur = conn.cursor()
    cur.execute("DELETE FROM pipeline_projects WHERE source_file = ?", (source_file,))
    deleted = cur.rowcount
    if deleted:
        print(f"  Cleared {deleted} prior rows for {source_file}")

    scanned = matched = 0
    for row in row_iter:
        scanned += 1
        rec = parse_row(row, hi)
        if not rec["_has_storage"]:
            continue
        if rec["iso_market"] not in TARGET_ISOS:
            continue
        matched += 1
        rec.pop("_has_storage")
        rec["source_file"] = source_file
        cur.execute(
            f"""
            INSERT INTO pipeline_projects
                ({', '.join(rec.keys())})
            VALUES ({', '.join('?' * len(rec))})
            """,
            list(rec.values()),
        )

    conn.commit()
    wb.close()
    print(f"  Scanned {scanned:,} rows, imported {matched:,} ERCOT/CAISO storage projects")
    return {"scanned": scanned, "matched": matched, "source_file": source_file}


def report_totals_by_iso(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    bar = "=" * 78
    print(f"\n{bar}\nTOTAL STORAGE CAPACITY BY ISO (gigawatts, nameplate power)\n{bar}")
    cur.execute(
        """
        SELECT iso_market, q_status, COUNT(*),
               ROUND(SUM(COALESCE(storage_mw, 0))/1000.0, 2)
        FROM pipeline_projects
        GROUP BY iso_market, q_status
        ORDER BY iso_market, 4 DESC
        """
    )
    last_iso = None
    iso_totals = defaultdict(float)
    iso_counts = defaultdict(int)
    for iso, status, n, gw in cur.fetchall():
        if iso != last_iso:
            if last_iso is not None:
                print(f"      {'─' * 12}")
                print(f"      Total: {iso_counts[last_iso]:>5} projects, "
                      f"{iso_totals[last_iso]:.2f} GW")
            print(f"\n  {iso}:")
            last_iso = iso
        iso_totals[iso] += gw or 0
        iso_counts[iso] += n
        print(f"    {status or '(unknown)':<14} {n:>5} projects   {gw or 0:>8.2f} GW")
    if last_iso:
        print(f"      {'─' * 12}")
        print(f"      Total: {iso_counts[last_iso]:>5} projects, {iso_totals[last_iso]:.2f} GW")

    print(f"\n  Note: GW = nameplate power capacity. The LBNL dataset has no duration column,")
    print(f"  so GWh (energy) cannot be computed directly. Multiply by 2-4h for an estimate.")


def report_top_developers(conn: sqlite3.Connection, n: int = 10) -> None:
    cur = conn.cursor()
    bar = "=" * 78
    print(f"\n{bar}\nTOP {n} DEVELOPERS BY STORAGE CAPACITY (ERCOT + CAISO)\n{bar}")
    cur.execute(
        """
        SELECT COALESCE(developer, '(not disclosed)') AS dev,
               COUNT(*) AS n_proj,
               ROUND(SUM(COALESCE(storage_mw, 0))/1000.0, 3) AS gw,
               SUM(CASE WHEN iso_market='ERCOT' THEN 1 ELSE 0 END) AS n_ercot,
               SUM(CASE WHEN iso_market='CAISO' THEN 1 ELSE 0 END) AS n_caiso
        FROM pipeline_projects
        WHERE storage_mw IS NOT NULL AND storage_mw > 0
        GROUP BY dev
        ORDER BY gw DESC
        LIMIT ?
        """,
        (n,),
    )
    print(f"  {'#':>2}  {'Developer':<40} {'Projects':>8}  {'GW':>7}  {'ERCOT/CAISO':>12}")
    print(f"  {'─' * 75}")
    for i, (dev, n_proj, gw, n_ercot, n_caiso) in enumerate(cur.fetchall(), 1):
        dev_disp = dev if len(dev) <= 40 else dev[:37] + "…"
        print(f"  {i:>2}  {dev_disp:<40} {n_proj:>8}  {gw or 0:>6.2f}   {n_ercot:>4}/{n_caiso:<4}")


def report_recent_filings(conn: sqlite3.Connection, days: int = 90) -> None:
    cur = conn.cursor()
    cutoff = (dt.date.today() - dt.timedelta(days=days)).isoformat()
    bar = "=" * 78
    print(f"\n{bar}\nPROJECTS FILED IN LAST {days} DAYS (since {cutoff})\n{bar}")
    cur.execute(
        """
        SELECT iso_market, q_date, q_status, state, developer, storage_mw, type_clean, project_name
        FROM pipeline_projects
        WHERE q_date >= ? AND q_date != ''
        ORDER BY q_date DESC, storage_mw DESC
        """,
        (cutoff,),
    )
    rows = cur.fetchall()
    if not rows:
        print("  (none — most recent LBNL release may predate the 90-day window)")
        cur.execute(
            "SELECT MAX(q_date) FROM pipeline_projects WHERE q_date IS NOT NULL AND q_date != ''"
        )
        max_date = cur.fetchone()[0]
        if max_date:
            print(f"  Latest q_date in the dataset: {max_date}")
        return

    print(f"  Found {len(rows)} project(s)\n")
    cur2 = conn.cursor()
    cur2.execute(
        """
        SELECT iso_market, COUNT(*), ROUND(SUM(COALESCE(storage_mw, 0))/1000.0, 2)
        FROM pipeline_projects WHERE q_date >= ? AND q_date != ''
        GROUP BY iso_market ORDER BY 3 DESC
        """,
        (cutoff,),
    )
    for iso, n, gw in cur2.fetchall():
        print(f"    {iso}: {n} projects, {gw or 0:.2f} GW")

    print(f"\n  Most recent filings (up to 20):\n")
    print(f"  {'Date':<11} {'ISO':<6} {'Status':<11} {'ST':<3} {'MW':>7}  Developer / Project")
    print(f"  {'─' * 75}")
    for iso, qd, st, state, dev, mw, tclean, name in rows[:20]:
        dev_label = dev or "(not disclosed)"
        proj_label = name or tclean or ""
        line = f"{dev_label}"
        if proj_label and proj_label != dev_label:
            line += f"  •  {proj_label}"
        print(f"  {qd:<11} {iso:<6} {st or '?':<11} {state or '?':<3} "
              f"{mw or 0:>6.0f}   {line[:55]}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("path", nargs="?", default=None,
                        help="Path to the LBNL Queued Up xlsx (default: auto-find latest in data/).")
    args = parser.parse_args()

    workbook_path = args.path or auto_locate_workbook()
    if not os.path.exists(workbook_path):
        sys.exit(f"Workbook not found: {workbook_path}")

    conn = init_db()
    stats = import_workbook(conn, workbook_path)

    if stats["matched"] == 0:
        print("\nNo matching ERCOT/CAISO storage projects found — nothing to report.")
        conn.close()
        return

    report_totals_by_iso(conn)
    report_top_developers(conn, n=10)
    report_recent_filings(conn, days=90)
    conn.close()


if __name__ == "__main__":
    main()
